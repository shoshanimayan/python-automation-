# sends email reminders based on payment status in spreadsheet
import openpyxl, smtplib, sys

# open the spreadsheet and get the latest dues status

wb = openpyxl.load_workbook('duesRecords.xlsx') #can change files here
sheet = wb.get_sheet_by_name('Sheet1')
lastCol = sheet.get_highest_column()
latestMonth = sheet.cell(row=1, column=lastCol).value

#check each members payment status
unpaid={}
for r in range(2,sheet.get_highest_row()+1):
    payment=sheet.cell(row=r, column=lastCol).value
    if payment !='paid':
        name= sheet.cell(row=r, column=1).value
        email = sheet.cell(row=r, column=2).value
        unpaid[name]=email

#log in to email account
smtpObj = smtplib.SMTP('stmp.gmail.com',587)
smtpOObj.ehlo()
smtpObj.starttls()
smtpObj.login('my_email_address@gamil.com', sys.argv[1]) #can change email

#send out email reminders
for name, email in unpaid.items():
    body="Subject: %s dues unpaid. \nDear %s, \nRecords show that you have not paid dues for %s. Please make this payment as soon as possible. Thank you'" %(lastestMonth, name, latestMonth)
    print('Sending email to %s...' % email)
    sendmailStatus= smtpObj.sendmail('my_email_address@gmail.com', email, body) #can change email

    if sendmailStatus !- {}:
        print('there was a problem sending  the email to: %s' (email))
smtpObj.quit()
